import boto3
import uuid
import pandas as pd
from io import BytesIO
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Dict
from app.core.settings import settings
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExportService:
    """
    Service responsible for exporting pandas DataFrames to Excel files.

    Supports two storage backends:
    - Local filesystem
    - AWS S3

    Storage backend is controlled via:
    - Constructor argument `storage_backend`
    - OR environment variable `STORAGE_BACKEND` (defaults to 'local')
    """

    def __init__(
        self,
        bucket_name: Optional[str] = None,
        storage_backend: Optional[str] = None,
        region: Optional[str] = None,
        local_export_dir: str = "exports",
    ):
        """
        Initialize the export service.
        """

        # Determine storage backend (env var takes fallback role)
        self.storage_backend = storage_backend or settings.storage_backend

        self.bucket_name = bucket_name or settings.bucket_name
        self.region = region or settings.region

        # Initialize S3 client only when using S3 backend
        if self.storage_backend == "s3":
            self.s3 = boto3.client("s3", region_name=region)
        else:
            self.s3 = None

        # Ensure local export directory exists
        self.local_export_dir = Path(local_export_dir)
        self.local_export_dir.mkdir(parents=True, exist_ok=True)

    # -------------------------
    # Public API
    # -------------------------
    def upload_excel(
        self,
        sheets: Dict[str, pd.DataFrame],
        prefix: str = "exports",
        filename_prefix: str = "export",
    ) -> str:
        """
        Export multiple DataFrames into a single Excel file.
        """

        if not sheets:
            raise ValueError("No sheets provided")

        for name, df in sheets.items():
            if df.empty:
                raise ValueError(f"Sheet '{name}' has no data")

        filename = self._generate_filename(filename_prefix)

        if self.storage_backend == "s3":
            return self._upload_to_s3(sheets, prefix, filename)
        else:
            return self._save_locally(sheets, prefix, filename)

    # -------------------------
    # Internal helpers
    # -------------------------
    def _generate_filename(self, prefix: str) -> str:
        unique_id = uuid.uuid4().hex
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}_{unique_id}.xlsx"

    def _apply_styles(self, writer, sheets: Dict[str, pd.DataFrame]):
        """
        Apply borders, auto-fit column widths, and header styling.
        """
        workbook = writer.book
        for sheet_name in sheets.keys():
            worksheet = workbook[sheet_name[:31]]

            # Define border style
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Apply borders to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

            # Style header row
            header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            for cell in worksheet[1]:  # First row = header
                cell.fill = header_fill
                cell.font = header_font

            # Auto-fit column widths
            for col in worksheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                worksheet.column_dimensions[col_letter].width = max_length + 2

    def _write_excel(self, sheets: Dict[str, pd.DataFrame]) -> BytesIO:
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            self._apply_styles(writer, sheets)
        buffer.seek(0)
        return buffer

    def _upload_to_s3(self, sheets: Dict[str, pd.DataFrame], prefix: str, filename: str) -> str:
        file_key = f"{prefix}/{filename}"
        buffer = self._write_excel(sheets)

        self.s3.put_object(
            Bucket=settings.bucket_name,
            Key=file_key,
            Body=buffer.getvalue(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return file_key

    def _save_locally(self, sheets: Dict[str, pd.DataFrame], prefix: str, filename: str) -> str:
        folder = self.local_export_dir / prefix
        folder.mkdir(parents=True, exist_ok=True)
        file_path = folder / filename

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            self._apply_styles(writer, sheets)

        return str(file_path)

    def generate_presigned_url(self, key: str, expires_in: int = 3600) -> str:
        if self.storage_backend == "s3":
            return self.s3.generate_presigned_url(
                ClientMethod="get_object",
                Params={"Bucket": self.bucket_name, "Key": key},
                ExpiresIn=expires_in,
            )
        return f"file:///{Path(key).resolve()}"
